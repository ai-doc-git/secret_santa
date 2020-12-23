import random
import copy
import smtplib
import openpyxl

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

names = []
emails = []
address = []
mobile = []

emp_data_file = "employee.xlsx"
wb = openpyxl.load_workbook(emp_data_file)
sheet = wb.active
max_row = sheet.max_row

for i in range(1, max_row + 1):
    names_t = sheet.cell(row = i, column = 1)
    emails_t = sheet.cell(row=i, column=2)
    address_t = sheet.cell(row=i, column=3)
    mobile_t = sheet.cell(row=i,column=4)
    names.append(names_t.value)
    emails.append(emails_t.value)
    address.append(address_t.value)
    mobile.append(mobile_t.value)

def secret_santa(names):
    my_list = names
    choose = copy.copy(my_list)
    result = []
    for i in my_list:
        names = copy.copy(my_list)
        names.pop(names.index(i))
        chosen = random.choice(list(set(choose)&set(names)))
        pos = my_list.index(chosen)
        result.append((i,chosen,address[pos],mobile[pos]))
        choose.pop(choose.index(chosen))
    return result

secret_santa_result = secret_santa(names)
final_package = zip(secret_santa_result,emails)

for x in final_package:
    fromaddr = "your email address"
    toaddr = x[1]
    msg = MIMEMultipart()
    msg['From'] = fromaddr
    msg['To'] = toaddr
    msg['Subject'] = "SECRET SANTA CELEBRATIONS!"

    body1 = "Hello, " + str(x[0][0])
    body2 = '''!\n It's time for our annual Secret Santa Gift Exchange.\n\nYou have been chosen as the secret santa of\n.......\n........\n........\n''' + str(
        x[0][1]) + "\nThere are some rules for this game, which are as follows : \n\nRule Number 1: Please don't tell anyone!\n"
    body3 = "Rule Number 2: The maximum budget for this year is Rs.500 \nWhat are you waiting for? Go ahead and get something nice for " + str(
        x[0][1]) + "!\n\n\n"
    body4 = "Please be a good Santa and send the gift as soon as possible to " + str(x[0][1]) + " at his address given below:\n\n " + str(x[0][2]) + "!\n\n\n"
    body5 = "Mobile number of " + str(x[0][1]) + " is " + str(x[0][3]) + "!\n\n\n"
    body = body1 + body2 + body3 + body4 + body5
    msg.attach(MIMEText(body, 'plain'))

    filename = "Merry Christmas"
    attachment = open("merry.jpg", "rb")

    part = MIMEBase('application', 'octet-stream')
    part.set_payload((attachment).read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= %s" % filename)

    msg.attach(part)

    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(fromaddr, "your password")
    text = msg.as_string()
    server.sendmail(fromaddr, toaddr, text)
    server.quit()
    print("mail sent to", x[1])